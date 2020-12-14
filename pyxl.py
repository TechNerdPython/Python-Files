from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('OpenPyXL')
root.geometry("500x800")


def get_a():
    list = ''
    for cell in col_a:
        list = f"{list + str(cell.value)}\n"

    label_a.config(text=list)


def get_b():
    list = ''
    for cell in col_b:
        list = f"{list + str(cell.value)}\n"

    label_b.config(text=list)


# Create workbook instance
wb = Workbook()

# Load existing workbook
wb = load_workbook('pizza.xlsx')

# Create active worksheet
ws = wb.active

# Create variable for column a
col_a = ws['A']
col_b = ws['B']

ba = Button(root, text="Get Column A", command=get_a)
ba.pack()

label_a = Label(root, text="")
label_a.pack(pady=20)

bb = Button(root, text="Get Column B", command=get_b)
bb.pack()

label_b = Label(root, text="")
label_b.pack(pady=20)

# Add Bob to A8 and B8
ws['A8'] = "Bob"
ws['B8'] = "Chicken"

# Save new file
wb.save('pizza2.xlsx')

root.mainloop()
