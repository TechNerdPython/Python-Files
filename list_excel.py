from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Lalith's Excel to Listbox")
root.geometry("400x300")

# Select Function
def select(e):
    my_label.config(text=my_list_box.get(ANCHOR))

# my_list = ["one", "two", "three"]

# Create Listbox
my_list_box = Listbox(root, width=45)
my_list_box.pack(pady=20)

# Create a Workbook
wb = load_workbook("name_color.xlsx")
# Set active worksheet
ws = wb.active

# Grab a column of data
col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_list_box.insert(END, item.value)

my_label = Label(root, text='Select an Item', font=("Helvetica", 18))
my_label.pack(pady=20)

# Create listbox binding
my_list_box.bind("<ButtonRelease-1>", select)

root.mainloop()